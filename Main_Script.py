import numpy as np
import pandas as pd
from openpyxl import *
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, LineChart, Reference
try:
    file = input("Paste the MyFitnessPal CSV file path here: ")
    df= pd.read_csv(file)
except:
    print("Not a Valid file path, make sure the file is in your directory and that the name is correct")

try:
    caloric_goal = int(input("How many calories do you aim for a day?: "))
except:
    print("Not a valid number")

cal_list = df.Calories
day_cals = []
cal_list2 = []
fats = [[],[],[],[]]
proteins = [[],[],[],[]]
carbs = [[],[],[],[]]
macros = ["fats","proteins","carbs"]
fats[0] = df["Fat (g)"].tolist()
proteins[0] = df["Protein (g)"].tolist()
carbs[0] = df["Carbohydrates (g)"].tolist()

for i in cal_list:
    cal_list2.append(i)


for k in range(0,(len(cal_list)//3)):
    k *= 3
    day_cals.append(cal_list2[k]+cal_list2[k+1]+cal_list2[k+2])

daily_cals = list(np.around(np.array(day_cals),2))

for i in range(0,(len(fats[0])//3)):
    i *= 3
    fats[1].append(fats[0][i]+fats[0][i+1]+fats[0][i+2])
for i in range(0,(len(proteins[0])//3)):
    i *= 3
    proteins[1].append(proteins[0][i]+proteins[0][i+1]+proteins[0][i+2])
for i in range(0,(len(carbs[0])//3)):
    i *= 3
    carbs[1].append(carbs[0][i]+carbs[0][i+1]+carbs[0][i+2])

max_macros = ["Max",max(daily_cals),max(carbs[1]),max(proteins[1]),max(fats[1])]
min_macros = ["Min",min(daily_cals),min(carbs[1]),min(proteins[1]),min(fats[1])]

print(min_macros)

alpha = {   "Calories": daily_cals,
            "Carbs": carbs[1],
            "Proteins": proteins[1],
            "Fats": fats[1]
}
days = []
#list of # of days
for i in range(0,len(fats[1])):
    days.append(f'Day {i+1}')

df2 = pd.DataFrame(data=alpha)
df2.insert(0,"Days",days)

df2_stats = pd.DataFrame(df2.describe())
avg_macros = ["Mean",
              round(df2_stats["Calories"].values[1],2),
              round(df2_stats["Carbs"].values[1],2),
              round(df2_stats["Proteins"].values[1],2),
              round(df2_stats["Fats"].values[1],2)]

print(df2_stats)
print(avg_macros)

row1 = pd.Series(max_macros, index=df2.columns)
df2 = df2._append(row1, ignore_index=True)
row2 = pd.Series(min_macros, index=df2.columns)
df2 = df2._append(row2, ignore_index=True)
row3 = pd.Series(avg_macros, index=df2.columns)
df2 = df2._append(row3, ignore_index=True)
print(df2)

pro_perc = []
carb_perc= []
fat_perc = []

for i in range(0,len(df2.Proteins)):
    pro_perc.append(round((df2["Proteins"].values[i]*4/df2["Calories"].values[i])*100,3))
for i in range(0,len(df2.Carbs)):
    carb_perc.append(round((df2["Carbs"].values[i]*4/df2["Calories"].values[i])*100,3))
for i in range(0,len(df2.Fats)):
    fat_perc.append(round((df2["Fats"].values[i]*9/df2["Calories"].values[i])*100,3))

df2.insert(5,"Percent Protein",pro_perc)
df2.insert(6,"Percent Carbs", carb_perc)
df2.insert(7,"Percent Fat", fat_perc)
print(df2)


df2.to_excel("New_Nutrition_Dataframe.xlsx")
wb = load_workbook("New_Nutrition_Dataframe.xlsx")
ws = wb.active

print(wb.worksheets)

def perc(num,avg):
    return(round(float((num-avg)/avg)*100,3))

perc_cals = []
for i in df2.Calories:
    perc_cals.append(perc(i,caloric_goal))

print(perc_cals)



#COLOR CODING
cella = ws["J3"]
lightred1 = PatternFill("solid", fgColor="ff9696")
lightred2 = PatternFill("solid", fgColor="ff3838")
redfill = PatternFill("solid", fgColor="c21b1b")
lightgreen1 = PatternFill("solid", fgColor="aaffa6")
lightgreen2 = PatternFill("solid", fgColor="6aff63")
greenfill = PatternFill("solid", fgColor="17e30e")

p_a = 1

for i in perc_cals:
    p_a += 1
    if i <= -20:
        ws[f"C{p_a}"].fill = redfill
    elif -20 < i <= -15:
        ws[f"C{p_a}"].fill = lightred2
    elif -15 < i <= -10:
        ws[f"C{p_a}"].fill = lightred1
    elif -10 < i <= -5:
        ws[f"C{p_a}"].fill = lightgreen2
    elif -5 < i <= 5:
        ws[f"C{p_a}"].fill = greenfill
    elif 5 < i <= 10:
        ws[f"C{p_a}"].fill = lightgreen1
    elif 10 < i <= 15:
        ws[f"C{p_a}"].fill = lightred1
    elif 15 < i <= 20:
        ws[f"C{p_a}"].fill = lightred2
    elif 20 < i <= 999:
        ws[f"C{p_a}"].fill = redfill
    else:
        pass

#BAR CHART OF EACH MACRO PER DAY
delta = len(df2.Calories) - 2
barchart1 = BarChart()
barchart1.type = "col"
barchart1.style = 2
barchart1.title = "Breakdown By Macro"
barchart1.y_axis.title = "# in Grams (g)"
barchart1.x_axis.title = "Day"
data1 = Reference(ws, min_col=4, max_col=6, min_row=1, max_row=delta)
barchart1.add_data(data1, titles_from_data=True)
ws.add_chart(barchart1,"M2")


linechart1 = LineChart()
linechart1.title = "Calories Per Day"
linechart1.style = 2
linechart1.y_axis.title = "Calories"
linechart1.x_axis.title = "Day"
data2 = Reference(ws, min_col=3, max_col=3, min_row=1, max_row=delta)
linechart1.add_data(data2,titles_from_data=True)
s1 = linechart1.series[0]
s1.marker.symbol = "square"
ws.add_chart(linechart1,"B19")

linechart2 = LineChart()
linechart2.title = "Macro % Breakdown Line Chart"
linechart2.style = 2
linechart2.y_axis.title = "macro percent (%)"
linechart2.x_axis.title = "Day"
data3 = Reference(ws, min_col=7, max_col=9, min_row=1, max_row=delta)
linechart2.add_data(data3,titles_from_data=True)
s2 = linechart2.series[0]
s2.marker.symbol = "square"
ws.add_chart(linechart2,"M19")



wb.save("New_Nutrition_Dataframe.xlsx")

